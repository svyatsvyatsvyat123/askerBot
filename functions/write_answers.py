import asyncio
import os
from time import time

from aiogram import Bot
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from bot_create import questions, answers_path


class Writer:
    wb: Workbook = None
    sheet: Worksheet = None
    bot: Bot = None
    last_ind: int = 1
    char_from_query: dict[str, str] = dict()
    path: list[str] = ['A' for _ in range(7)]
    sz: int = 0
    file_name: list[str] = ['A' for _ in range(7)]
    last_save_time = 0

    @staticmethod
    def next(s: list[str]) -> list[str]:
        c = s.copy()
        for i in range(len(c) - 1, -1, -1):
            if c[i] != 'Z':
                c[i] = chr(ord(c[i]) + 1)
                for j in range(i + 1, len(c)):
                    c[j] = 'A'
                return c
        return ['A' for i in range(len(c) + 1)]

    @staticmethod
    def make_dirs(path: list[str]) -> list[str]:
        while True:
            try:
                os.makedirs(f"{answers_path}/" + ''.join(path))
                return path
            except FileExistsError:
                file_cnt = len([_ for _ in os.listdir(f"{answers_path}/" + ''.join(path))])
                if file_cnt == 0:
                    return path
                if file_cnt < 50:
                    last = sorted([x for x in os.listdir(f"{answers_path}/" + ''.join(path))])[-1]
                    Writer.file_name = Writer.next(list(last.split(".")[0]))
                    return path
                path = Writer.next(path)

    @staticmethod
    async def open(bot: Bot, delete_answers=False):
        Writer.bot = bot
        os.makedirs(f"{answers_path}", exist_ok=True)
        try:
            if delete_answers:
                Writer.wb = Workbook()
            else:
                Writer.wb = load_workbook(f"{answers_path}/answers.xlsx")
                print(f"Файл {answers_path}/answers.xlsx обнаружен.")
        except FileNotFoundError:
            print(f"Файл {answers_path}/answers.xlsx не обнаружен и будет создан автоматически")
            Writer.wb = Workbook()
        try:
            Writer.sheet = Writer.wb.get_sheet_by_name("Sheet")
        except KeyError:
            if not delete_answers:
                print(f"Лист Sheet не обнаружен и будет создан автоматически")
            Writer.wb.create_sheet("Sheet")
            Writer.sheet = Writer.wb.get_sheet_by_name("Sheet")

        c = ['A']
        for i in range(len(questions)):
            if questions[i]["name"] is None:
                continue
            if Writer.sheet[f"{''.join(c)}1"].value is None:
                Writer.sheet[f"{''.join(c)}1"] = questions[i]["name"]
            elif Writer.sheet[f"{''.join(c)}1"].value != questions[i]["name"]:
                res = input("В существуещем файле answers.xlsx другая разметка столбцов."
                            "Введите (Y,yes,д,да) в любом регистре если хотите ее ОЧИСТИТЬ, или (все остальное),"
                            " если завершить программу:\n").lower().strip()
                if res in ("да", "y", "yes", "д"):
                    await Writer.open(bot, True)
                    return
                else:
                    Writer.wb.close()
                    exit(0)
            Writer.char_from_query[questions[i]["name"]] = ''.join(c)
            c = Writer.next(c)
        for i in range(20):
            if Writer.sheet[f"{''.join(c)}1"].value is None:
                Writer.sheet[f"{''.join(c)}1"] = f"Ф{i + 1}"
            elif Writer.sheet[f"{''.join(c)}1"].value != f"Ф{i + 1}":
                res = input("В существуещем файле answers.xlsx другая разметка столбцов."
                            "Введите (Y,yes,д,да) в любом регистре, если хотите ее ПЕРЕЗАПИСАТЬ, или (все остальное),"
                            " чтобы завершить программу:\n").lower().strip()
                if res in ("да", "y", "yes", "д"):
                    await Writer.open(bot, True)
                    return
                else:
                    Writer.wb.close()
                    exit(0)
            Writer.sheet[f"{''.join(c)}1"] = f"Ф{i + 1}"
            Writer.char_from_query[f"Ф{i + 1}"] = ''.join(c)
            c = Writer.next(c)
        Writer.wb.save(f"{answers_path}/answers.xlsx")
        Writer.path = Writer.make_dirs(Writer.path)
        if delete_answers:
            print("Таблица успешно отфарматирована")
        else:
            i = 2
            while Writer.sheet[f"A{i}"].value is not None:
                Writer.last_ind += 1
                i += 1

    @staticmethod
    async def write(data: dict[str, str]):
        Writer.last_ind += 1
        for question in questions:
            name: str = question["name"]
            if name is not None:
                Writer.sheet[f"{Writer.char_from_query[name]}{Writer.last_ind}"] = data[name]
        if Writer.sz + int(data["cnt"]) > 50:
            Writer.path = Writer.next(Writer.path)
            Writer.sz = 0
            Writer.path = Writer.make_dirs(Writer.path)
        path = ''.join(Writer.path)
        Writer.sz += int(data["cnt"])
        file_name = Writer.file_name
        file_names = [''.join(Writer.file_name)]
        for i in range(int(data["cnt"]) - 1):
            file_name = Writer.next(file_name)
            file_names.append(''.join(file_name))
        Writer.file_name = Writer.next(file_name)
        file_types = data['files_names'].split()
        files_id = data['files_id'].split()
        files_all_path = [f"{answers_path}/" + path + r"/" + file_names[i] + '.' + file_types[i] for i in
                          range(len(file_types))]
        tasks = (
            Writer.bot.download_file_by_id(files_id[i], files_all_path[i])
            for i in range(len(file_types)))
        for i in range(len(file_types)):
            Writer.sheet[Writer.char_from_query[f"Ф{i + 1}"] + f"{Writer.last_ind}"] = files_all_path[i]
        await asyncio.gather(*tasks)
        if time() - Writer.last_save_time > 300:
            Writer.last_save_time = time()
            Writer.wb.save(f"{answers_path}/answers.xlsx")
            print("Новые данные успешно добавлены в таблицу")

    @staticmethod
    async def close(*args, **kwargs):
        if Writer.wb is not None:
            Writer.wb.save(f"{answers_path}/answers.xlsx")
            Writer.wb.close()
            Writer.wb = None
